"""Layout discovery in the workbook reader.

Backfilling made this worth testing on its own. The reader was written against one
workbook and quietly assumed its geometry; a second workbook with a different number of
columns then read the wrong cells rather than failing, which is the hardest kind of wrong
to notice in a reconciliation whose whole job is to notice things.
"""

import datetime as dt
from decimal import Decimal

import openpyxl
import pytest

from budget import xlsm_reader as xr


def make_summary(card_columns: int) -> openpyxl.Workbook:
    """A Summary sheet shaped like the real one: a credit card table, then the
    month-by-classification block to its right. The card count is what moves the block --
    25-26 has three cards where 26-27 has five, which shifts it two columns left."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    col = 9  # I, where the card table starts in both workbooks
    ws.cell(19, col, "Month")
    for i in range(card_columns):
        ws.cell(19, col + 1 + i, f"Card {i + 1}")
    ws.cell(19, col + 1 + card_columns, "Payment")
    for i, month in enumerate(xr.FISCAL_MONTHS):
        ws.cell(20 + i, col, month)

    col += card_columns + 3
    ws.cell(19, col, "Month")
    for i, name in enumerate(["Bills", "Excess", "Food", "Spend"]):
        ws.cell(19, col + 1 + i, name)
    for i, month in enumerate(xr.FISCAL_MONTHS):
        ws.cell(20 + i, col, month)
        ws.cell(20 + i, col + 1, 100 + i)

    return wb, col


@pytest.fixture
def ref():
    return xr.RefData(
        accounts=[],
        categories=[],
        classifications=[
            xr.ClassificationRef(name=n, legacy_ref=i, direction=1, display_order=i)
            for i, n in enumerate(["Bills", "Excess", "Food"])
        ],
        months={m: (i + 4 - 1) % 12 + 1 for i, m in enumerate(xr.FISCAL_MONTHS)},
        settings={"tax_year": "2025"},
    )


@pytest.mark.parametrize("cards", [3, 5])
def test_finds_the_block_wherever_the_card_table_leaves_it(ref, cards):
    wb, expected_col = make_summary(cards)

    month_col, headers, rows = xr.summary_matrix(wb, ref)

    assert month_col == expected_col
    assert list(headers.values()) == ["Bills", "Excess", "Food", "Spend"]
    assert list(rows) == list(range(20, 32))


def test_skips_the_credit_card_table(ref):
    """Both tables head their first column 'Month'. Picking the left one is what read a
    card balance as a month name and asked period_for() about '-4123.37'."""
    wb, expected_col = make_summary(3)

    month_col, _, _ = xr.summary_matrix(wb, ref)

    assert month_col != 9
    assert month_col == expected_col


def test_says_so_when_there_is_no_block(ref):
    wb = openpyxl.Workbook()
    wb.active.title = "Summary"

    with pytest.raises(ValueError, match="no month-by-classification block"):
        xr.summary_matrix(wb, ref)


# --------------------------------------------------------------------------- salary bands


def make_salary_tracker(additional_rate: bool) -> openpyxl.Workbook:
    """The band block. 25-26 models no additional rate, so everything below it sits one
    row higher than in 26-27 -- which is what a row-based read cannot survive."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary tracker"

    rows = [
        ("NI", "LEL", 1048, None),
        (None, "UEL", 4189, None),
        (None, "Lower rate", 0.08, None),
        (None, "Higher rate", 0.02, None),
        ("PAYE", "Personal allowance", 1047.5, None),
        (None, None, None, None),
        (None, "PA - 1", -55, dt.date(2026, 4, 1)),
        (None, "PA - 2", -246.42, dt.date(2026, 6, 1)),
        (None, "Basic rate threshold", 4189.17, None),
        (None, None, None, None),
        (None, "Higher rate threshold", 10428.33, None),
        (None, None, None, None),
        (None, "Basic rate", 0.2, None),
        (None, "Higher rate", 0.4, None),
    ]
    if additional_rate:
        rows.append((None, "Additional rate", 0.45, None))
    rows += [
        ("Adjusted bands", "Personal allowance", -55, None),
        (None, "Basic rate", 3141.67, None),
    ]

    for offset, (section, band, value, start) in enumerate(rows):
        row = 18 + offset
        if section:
            ws.cell(row, 2, section)
        if band:
            ws.cell(row, 3, band)
        if value is not None:
            ws.cell(row, 4, value)
        if start:
            ws.cell(row, 6, start)
    return wb


def test_bands_are_found_by_name_not_by_row():
    """Both shapes read correctly, and the block moving does not shift what is read."""
    with_extra = dict(
        (k, v) for k, _, v in xr.read_salary_assumptions(make_salary_tracker(True), 2026)
    )
    without = dict(
        (k, v) for k, _, v in xr.read_salary_assumptions(make_salary_tracker(False), 2025)
    )

    assert with_extra["additional_rate"] == Decimal("45.00")
    assert with_extra["basic_band"] == Decimal("3141.67")
    # Absent, not zero: a workbook that models no additional rate has not said it is 0%.
    assert "additional_rate" not in without
    assert without["basic_band"] == Decimal("3141.67")
    assert without["basic_rate"] == Decimal("20.00")


def test_the_adjusted_personal_allowance_is_not_read_as_a_rate():
    """The failure this replaced: 'Adjusted bands -> Personal allowance' sat on the row a
    row-based read expected 'Additional rate' on, storing 475.83 as 47,583%."""
    bands = dict((k, v) for k, _, v in xr.read_salary_assumptions(make_salary_tracker(False), 2025))

    assert bands["personal_allowance"] == Decimal("1047.5")
    assert all(value < Decimal("100") for key, value in bands.items() if key.endswith("_rate"))


def test_allowance_steps_keep_their_own_dates():
    steps = [
        (start, value)
        for key, start, value in xr.read_salary_assumptions(make_salary_tracker(True), 2026)
        if key == "personal_allowance_adjustment"
    ]

    assert steps == [(dt.date(2026, 4, 1), Decimal("-55")), (dt.date(2026, 6, 1), Decimal("-246.42"))]


# --------------------------------------------------------------------------------- cards


def make_cards(names: list[str]) -> openpyxl.Workbook:
    """Two columns of schedule per card, then the parameter block -- so the block starts
    further right the more cards there are."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Transfer Cards"

    col = 3
    for name in names:
        ws.cell(2, col, name)
        col += 2
    ws.cell(2, col, "Total")
    col += 2

    ws.cell(2, col + 1, "Payment dates")
    ws.cell(2, col + 2, "Term (months)")
    ws.cell(2, col + 3, "Min payment")
    for i, name in enumerate(names):
        ws.cell(3 + i, col, name)
        ws.cell(3 + i, col + 1, 5)
        ws.cell(3 + i, col + 2, 10 + i)
        ws.cell(3 + i, col + 3, 0.01)

    ws.cell(4, 2, dt.date(2025, 4, 1))
    for i, _ in enumerate(names):
        ws.cell(4, 3 + i * 2, 1000 + i)
    return wb


@pytest.mark.parametrize("names", [["A", "B", "C"], ["A", "B", "C", "D", "E"]])
def test_cards_are_found_wherever_the_schedule_ends(names):
    cards = xr.read_cards(make_cards(names))

    assert [c["name"] for c in cards] == names
    assert [c["term_months"] for c in cards] == [10 + i for i in range(len(names))]
    assert cards[0]["opening_balance"] == Decimal("1000")
    assert cards[0]["opening_date"] == dt.date(2025, 4, 1)


def test_cards_say_so_when_the_block_is_missing():
    wb = openpyxl.Workbook()
    wb.active.title = "Balance Transfer Cards"

    with pytest.raises(ValueError, match="Payment dates"):
        xr.read_cards(wb)
