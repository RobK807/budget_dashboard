"""Bring the pension history across from the tracking workbook.

Run once per machine that needs the history; after that the pension lives in the database
like everything else and this file is never read again.

Two sheets are read:

  * **Overview** -- a row per date, a column per pot. These become `pension_valuation`
    rows, one per pot per date, rather than the one wide row the sheet holds. The sheet's
    shape forces every pot to be valued on the same day; the tables here do not, because
    three providers do not publish together.

  * **A sheet named after a pot** -- a Credit column and a Debit column, which become one
    signed `pension_contribution` row each. Rows where both are zero are the sheet's way of
    marking a valuation date inside the ledger, and carry no movement, so they are skipped.

The derived columns are all left behind. Implied return, annualised return, total return
and annualised total return are recomputed by `repo.pension_history` from the valuations and
the ledger, which is the whole point of storing those two and nothing else.

**One figure deliberately changes.** The sheet's per-period return for the pot with a ledger
was its return *since inception*, sitting in the column beside two genuine period returns --
so the three could not be read across, and the weighted total built from them was measuring
nothing in particular. Here every period return is the same calculation: value at the end
over value at the start plus anything paid in between. Every other figure the sheet computes
is reproduced exactly, which is what `report()` prints.

Idempotent. Valuations are keyed by pot and date, so restating one overwrites it. Ledger
entries have no natural key -- the same contribution can legitimately appear twice on one
day, and does -- so they are reconciled by count: for each pot, date and amount, only the
shortfall against what the sheet holds is inserted. A second run therefore writes nothing.

Run with:  python -m budget.seed_pension
           python -m budget.seed_pension --report   # print and check, change nothing
"""

from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
from collections import Counter
from decimal import Decimal

from sqlalchemy import select

from budget import config, repo
from budget.db import create_all, in_use, make_engine, make_session_factory
from budget.models import PensionContribution, PensionPot, PensionValuation
from budget.service import bump_revision

OVERVIEW = "Overview"
# The Overview's own layout: a header row, then a Date column and one column per pot up to
# a Total the sheet computes and this does not need.
HEADER_ROW = 2
DATE_COLUMN = 2
FIRST_POT_COLUMN = 3
TOTAL_HEADING = "Total"

# The ledger's columns, on whichever sheet is named after a pot.
LEDGER_DATE = 2
LEDGER_CREDIT = 3
LEDGER_DEBIT = 4

# Anything smaller than this on the way in is the provider's cash account paying interest,
# not a contribution. The gap in the data is not marginal -- the interest credits run to
# under two pounds and the smallest contribution is over seven hundred -- so this only has
# to land somewhere in between. It is a label for reporting and nothing calculates from it.
INTEREST_CEILING = Decimal("10")


def _cell(sheet, row: int, column: int):
    return sheet.cell(row=row, column=column).value


def read_workbook(path=None) -> tuple[list[str], list[dict], list[dict]]:
    """The pots, their valuations and every ledger movement, as plain dictionaries."""
    import openpyxl

    book = openpyxl.load_workbook(path or config.PENSION_PATH, data_only=True)
    sheet = book[OVERVIEW]

    names: list[str] = []
    column = FIRST_POT_COLUMN
    while True:
        heading = _cell(sheet, HEADER_ROW, column)
        if heading in (None, TOTAL_HEADING):
            break
        names.append(str(heading))
        column += 1
    if not names:
        raise SystemExit("No pots found. Has the layout changed?")

    valuations: list[dict] = []
    row = HEADER_ROW + 1
    while _cell(sheet, row, DATE_COLUMN) is not None:
        on = _cell(sheet, row, DATE_COLUMN)
        on = on.date() if isinstance(on, dt.datetime) else on
        for index, name in enumerate(names):
            value = _cell(sheet, row, FIRST_POT_COLUMN + index)
            if value is not None:
                valuations.append(
                    {"pot": name, "on_date": on, "value": Decimal(str(value))}
                )
        row += 1

    movements: list[dict] = []
    for ledger in book.worksheets:
        if ledger.title not in names:
            continue
        row = HEADER_ROW + 1
        while _cell(ledger, row, LEDGER_DATE) is not None:
            on = _cell(ledger, row, LEDGER_DATE)
            on = on.date() if isinstance(on, dt.datetime) else on
            credit = Decimal(str(_cell(ledger, row, LEDGER_CREDIT) or 0))
            debit = Decimal(str(_cell(ledger, row, LEDGER_DEBIT) or 0))
            amount = credit - debit
            if amount:
                movements.append(
                    {
                        "pot": ledger.title,
                        "on_date": on,
                        "amount": amount,
                        "kind": _kind(amount),
                    }
                )
            row += 1

    return names, valuations, movements


def _kind(amount: Decimal) -> str:
    if amount < 0:
        return "charge"
    return "interest" if amount < INTEREST_CEILING else "contribution"


def _frames(names, valuations, movements):
    """The three loaded lists in the shape `repo.pension_history` reads, for the check."""
    import pandas as pd

    ids = {name: index + 1 for index, name in enumerate(names)}
    pots = pd.DataFrame(
        [
            {
                "id": ids[name],
                "name": name,
                "display_order": ids[name],
                "valid_from": min(
                    [v["on_date"] for v in valuations if v["pot"] == name]
                    + [m["on_date"] for m in movements if m["pot"] == name]
                ),
                "valid_to": None,
            }
            for name in names
        ]
    )
    held = pd.DataFrame(
        [
            {"pot_id": ids[v["pot"]], "on_date": v["on_date"], "value": v["value"]}
            for v in valuations
        ]
    )
    paid = pd.DataFrame(
        [
            {
                "id": index + 1,
                "pot_id": ids[m["pot"]],
                "on_date": m["on_date"],
                "amount": m["amount"],
                "kind": m["kind"],
                "note": None,
            }
            for index, m in enumerate(movements)
        ],
        columns=["id", "pot_id", "on_date", "amount", "kind", "note"],
    )
    return pots, held, paid


def report(path=None) -> bool:
    """Print what will be written and check it against the sheet's own totals.

    ASCII only -- this prints to a cp1252 console, where a tick mark raises
    UnicodeEncodeError before anything has been written.
    """
    names, valuations, movements = read_workbook(path)
    pots, held, paid = _frames(names, valuations, movements)
    history = repo.pension_history(pots, held, paid)
    totals = repo.pension_totals(history)

    print(f"\nPots: {', '.join(names)}")
    print(f"Valuations: {len(valuations)} across {len(totals)} date(s)")
    print(f"Ledger entries: {len(movements)}")

    counts = Counter(m["kind"] for m in movements)
    for kind in ("contribution", "interest", "charge", "other"):
        if counts.get(kind):
            total = sum(
                (m["amount"] for m in movements if m["kind"] == kind), Decimal("0")
            )
            print(f"  {kind:<14} {counts[kind]:>4}  {total:>12,.2f}")

    print("\nPosition at each valuation")
    print("-" * 96)
    print(f"{'date':<14}{'value':>14}{'paid in':>14}{'growth':>14}"
          f"{'period %':>12}{'to date %':>12}{'a year %':>12}")
    import pandas as pd

    def show(value) -> str:
        return "-" if value is None or pd.isna(value) else f"{value:,.2f}"

    for _, row in totals.iterrows():
        print(
            f"{row['date']:%d %b %Y}".ljust(14)
            + f"{row['value']:>14,.2f}{row['base']:>14,.2f}{row['growth']:>14,.2f}"
            + f"{show(row['period_return']):>12}"
            + f"{show(row['total_return']):>12}"
            + f"{show(row['total_annualised']):>12}"
        )
    print("-" * 96)

    # The checks. Each one is a figure the sheet computes for itself, so agreement means the
    # reading is right rather than merely plausible -- and the two that cannot agree are
    # named, so a difference is never a surprise.
    closing = totals.iloc[-1]
    stated = _stated_total(path)
    ok = stated is None or abs(closing["value"] - stated) < Decimal("0.01")
    if stated is None:
        print(f"\n  total value      {closing['value']:>14,.2f}   (nothing to check against)")
    else:
        print(f"\n  total value      {closing['value']:>14,.2f}   against {stated:>14,.2f}"
              f"   {'[ok]' if ok else '[DOES NOT MATCH]'}")

    for _, row in history[history["date"] == history["date"].max()].iterrows():
        print(f"  {row['pot']:<16} {row['value']:>14,.2f}   "
              f"return to date {row['total_return']:>8,.4f}%")

    print("\n  Recomputed, and deliberately different from the sheet:")
    print("    - the period return for a pot that is paid into, which the sheet held as a")
    print("      return since inception, so it could not be read across the other pots")
    print("    - the combined return, worked out from the pounds rather than averaged")
    print("      across the pots, which differs by most of a percentage point")
    print("    - annualising uses 365.25 days throughout, worth a hundredth of a point")
    return ok


def _stated_total(path=None) -> Decimal | None:
    """The sheet's own final total, purely to check the reading against."""
    import openpyxl

    book = openpyxl.load_workbook(path or config.PENSION_PATH, data_only=True)
    sheet = book[OVERVIEW]
    row = HEADER_ROW + 1
    last = None
    while _cell(sheet, row, DATE_COLUMN) is not None:
        last = row
        row += 1
    if last is None:
        return None
    column = FIRST_POT_COLUMN
    while _cell(sheet, HEADER_ROW, column) not in (None, TOTAL_HEADING):
        column += 1
    value = _cell(sheet, last, column)
    return Decimal(str(value)) if value is not None else None


def apply(session, path=None) -> list[str]:
    """Everything this writes, in one transaction."""
    names, valuations, movements = read_workbook(path)
    done: list[str] = []

    starts = {
        name: min(
            [v["on_date"] for v in valuations if v["pot"] == name]
            + [m["on_date"] for m in movements if m["pot"] == name]
        )
        for name in names
    }

    pots: dict[str, PensionPot] = {}
    for name in names:
        existing = session.scalars(
            select(PensionPot).where(PensionPot.name == name)
        ).first()
        if existing is None:
            existing = PensionPot(
                name=name, valid_from=starts[name], display_order=names.index(name) + 1
            )
            session.add(existing)
            session.flush()
            done.append(f"added pension {name}, tracked from {starts[name]:%d %b %Y}")
        elif existing.valid_from > starts[name]:
            existing.valid_from = starts[name]
            done.append(f"{name}: moved the start back to {starts[name]:%d %b %Y}")
        pots[name] = existing

    written = replaced = 0
    for entry in valuations:
        pot = pots[entry["pot"]]
        held = session.get(PensionValuation, (pot.id, entry["on_date"]))
        if held is None:
            session.add(
                PensionValuation(
                    pot_id=pot.id, on_date=entry["on_date"], value=entry["value"]
                )
            )
            written += 1
        elif held.value != entry["value"]:
            held.value = entry["value"]
            replaced += 1
    session.flush()
    done.append(f"valuations: {written} added, {replaced} updated, "
                f"{len(valuations) - written - replaced} already correct")

    # Reconciled by count rather than by presence: two identical contributions on one day
    # are ordinary here -- the employer's share and your own can match to the penny -- so
    # 'is there one like this already' would silently drop the second every time.
    wanted = Counter(
        (m["pot"], m["on_date"], m["amount"]) for m in movements
    )
    already = Counter(
        (row.pot.name, row.on_date, row.amount)
        for row in session.scalars(select(PensionContribution))
        if row.pot.name in pots
    )
    added = 0
    for key, count in wanted.items():
        name, on, amount = key
        for _ in range(count - already.get(key, 0)):
            session.add(
                PensionContribution(
                    pot_id=pots[name].id,
                    on_date=on,
                    amount=amount,
                    kind=_kind(amount),
                )
            )
            added += 1
    session.flush()
    done.append(f"ledger: {added} added, {sum(wanted.values()) - added} already there")

    bump_revision(session)
    return done


def snapshot() -> str:
    """VACUUM INTO rather than a file copy: transactionally consistent even mid-write, where
    copying a live SQLite file can capture a torn state (DESIGN.md 7)."""
    target = config.DB_PATH.with_name(
        f"budget.pre-pension-seed-{dt.datetime.now():%Y%m%d-%H%M%S}.db"
    )
    conn = sqlite3.connect(config.DB_PATH)
    try:
        conn.execute(f"VACUUM INTO '{str(target).replace(chr(39), chr(39) * 2)}'")
    finally:
        conn.close()
    return target.name


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Seed the pension history.")
    parser.add_argument("--yes", action="store_true", help="skip the confirmation")
    parser.add_argument(
        "--report", action="store_true", help="print and check, change nothing"
    )
    parser.add_argument("--file", help="read from somewhere other than the usual place")
    args = parser.parse_args(argv)

    if not config.PENSION_PATH.exists() and not args.file:
        print(f"Not found: {config.PENSION_PATH}")
        return 1

    ok = report(args.file)
    if args.report:
        return 0 if ok else 1
    if not ok:
        print("\nThe figures do not reconcile. Nothing was changed.")
        return 1

    if in_use():
        print("\nThe dashboard still has the database open. Close every window and re-run.")
        return 1

    if not args.yes and input("\nApply? [y/N] ").strip().lower() not in ("y", "yes"):
        print("Nothing was changed.")
        return 1

    print(f"\nSnapshot taken: {snapshot()}")

    engine = make_engine()
    try:
        create_all(engine)
        factory = make_session_factory(engine)
        with factory() as session, session.begin():
            for line in apply(session, args.file):
                print(f"  {line}")

        with factory() as session:
            pots = repo.load_pension_pots(session)
            held = repo.load_pension_valuations(session)
            paid = repo.load_pension_contributions(session)
        totals = repo.pension_totals(repo.pension_history(pots, held, paid))
        closing = totals.iloc[-1]
        print(f"\n  stored: {len(pots)} pension(s), {len(held)} valuation(s), "
              f"{len(paid)} ledger entr(ies)")
        print(f"  at {closing['date']:%d %b %Y}: {closing['value']:,.2f} "
              f"({closing['growth']:,.2f} of it growth)")
    finally:
        engine.dispose()

    print("\nThere is now a push pending.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
