"""Reconciles the migrated database against the workbook.

This is the Phase 0 acceptance gate (DESIGN.md 4, step 6). Two independent checks:

A. Account balances -- opening balance plus every signed posting must reproduce the
   month tab's row-61 'End' figure, for every account in every month.

B. Daily classification totals -- the per-day, per-classification figures the workbook
   computes with ~44 SUMIFS per cell must match a GROUP BY over the ledger.

A passes only if the posting rules and account types are right; B passes only if the
classification sign convention is right. They fail in different ways, which is why both
are worth running.

Run with:  python -m budget.reconcile [--db PATH] [--workbook PATH] [-v]
"""

from __future__ import annotations

import argparse
import datetime as dt
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl.utils import get_column_letter
from sqlalchemy import select

from budget import config, repo, xlsm_reader as xr
from budget.backfill_year import db_accounts_for_block
from budget.db import make_engine, make_session_factory
from budget.models import Account, Classification, OpeningBalance, Txn
from budget.postings import postings_for

TOLERANCE = Decimal("0.01")

# Differences where the database is deliberately right and the workbook is wrong. Each one
# is investigated and written up in PHASE0_FINDINGS.md before it is allowed in here -- the
# point of the gate is to stay green so that a genuine regression is visible, which it
# cannot be if the run is permanently red.
#
# Held per workbook, for the same reason Corrections is: (month, key) is not unique across
# years. 26-27 accepts ('May', 'Food') for a GBP 17.09 date fix, and while these shared one
# dict that entry silently swallowed an unrelated GBP 4.75 difference in 25-26's May -- an
# acceptance list that hides a difference it was never shown is worse than no list at all.
ACCEPTED_26_27: dict[tuple[str, str], str] = {
    ("May", "Food 2026-05-29"): (
        "Debug row 360 was dated 2029-05-29, so the workbook's MATCH found no day to post "
        "it against and dropped GBP 17.09 of Food spending. The corrected date restores it."
    ),
    ("June", "Going Out [Income]"): (
        "txn 566, 30 Jun, Platinum Amex, GBP 100, 'Cashback'. The month tab files this under "
        "Going Out; the ledger says Other, which is correct -- the 27 Jun cashback of GBP 100 "
        "is Other in both. Confirmed by the user; the workbook is wrong here."
    ),
    ("June", "Other [Income]"): "Counterpart of Going Out [Income]: the same GBP 100.",
    # The same GBP 17.09, seen again downstream: the running total is cumulative, so a
    # corrected day shifts the month's closing figure and everything that reads it.
    ("May", "Food"): "Summary view of the GBP 17.09 correction above.",
    ("May", "Running Food"): "Month-tab running view of the GBP 17.09 correction above.",
}

# Budget 25-26.xlsm. Two shapes recur, and both come of the month tabs being maintained by
# a macro rather than by formula: a row the tab put in the wrong month or at the wrong
# amount, and a row whose purchase type or category the tab records differently from the
# ledger. In every case below the ledger is right, confirmed by the user.
ACCEPTED_25_26: dict[tuple[str, str], str] = {
    ("March", "Amex"): (
        "Debug row 1929, 29 Mar, the NY hotel charge. The workbook says GBP 500.61; the real "
        "figure is 499.85, which the user identified. The 76p is exactly what stopped "
        "Platinum Amex's opening reconciling into 26-27, and with it corrected the two cards "
        "land on their stated April 2026 openings to the penny."
    ),
    ("February", "Mastercard"): (
        "Debug row 1796, 27 Feb, Mastercard, GBP 85.05 'Lottery'. The workbook posted it to "
        "the March tab; the ledger dates it correctly. Confirmed by the user."
    ),
    # Seven rows where the month tab's purchase type differs from the ledger's. The ledger
    # is right in each -- it agrees with how the same thing is recorded everywhere else in
    # the year, and the tab is the outlier.
    #
    # Each shows up twice, once under the type the tab used and once under the type it did
    # not, so both halves are listed. Balances are unaffected: the money is on the tab, only
    # its purchase type is not.
    ("May", "Excess 2025-05-01"): (
        "Debug row 194, 1 May, Amex, GBP 3.20 'Coffee'. The ledger says Excess; the tab "
        "says Food. Every other Coffee entry in the year is Excess."
    ),
    ("May", "Food 2025-05-01"): "Counterpart of Excess 2025-05-01: the same GBP 3.20.",
    ("May", "Excess 2025-05-02"): (
        "Debug row 190, 2 May, Amex, GBP 7.95 'Waitrose'. The ledger says Food; the tab "
        "says Excess. Every other Waitrose entry in the year is Food. Rows 190 and 194 sit "
        "out of date order, which is what a pair entered later, and crossed, looks like."
    ),
    ("May", "Food 2025-05-02"): "Counterpart of Excess 2025-05-02: the same GBP 7.95.",
    # These three carry no purchase type on the tab at all, so it counts nothing for them
    # and there is no second half to list.
    ("May", "Excess 2025-05-10"): (
        "Debug row 250, 10 May, Savings - Spending, GBP 5.00 interest. Excess in the "
        "ledger; the tab gives it no purchase type. The balance itself reconciles."
    ),
    ("May", "Excess 2025-05-30"): (
        "Debug row 331, 30 May, Savings - Marcus, GBP 18.85 interest. As 10 May."
    ),
    ("June", "Bills 2025-06-06"): (
        "Debug row 373, 6 June, Halifax, GBP 12.25 'Lottery'. The ledger says Bills; the "
        "tab says Excess. The other Halifax Lottery entries -- 9 May, 9 January -- are "
        "Bills and reconcile, so this one is the outlier."
    ),
    ("June", "Excess 2025-06-06"): "Counterpart of Bills 2025-06-06: the same GBP 12.25.",
    ("July", "Bills 2025-07-07"): (
        "Debug row 564, 7 July, Halifax, GBP 19.50 'Phone'. The ledger says Bills; the tab "
        "says Excess."
    ),
    ("July", "Excess 2025-07-07"): "Counterpart of Bills 2025-07-07: the same GBP 19.50.",
    ("July", "Excess 2025-07-10"): (
        "Debug row 598, 10 July, Savings - Spending, GBP 5.36 interest. As 10 May."
    ),
    # The daily view of two differences already accepted above.
    ("February", "Excess 2026-02-27"): (
        "Daily view of the February Mastercard difference: Debug row 1796's GBP 85.05, "
        "which the workbook posted to the March tab."
    ),
    ("March", "Savings 2026-03-29"): (
        "Daily view of the March Amex difference: Debug row 1929 at 499.85, not 500.61."
    ),
    # The category accumulator drifting one row. 'Expenses' sits directly above 'Other' in
    # the month tab's category block, and for a run of entries the macro added an 'Other'
    # amount to the Expenses row instead.
    #
    # These are accumulators, not formulas (New_entry does `ActiveCell = dblAmt +
    # dblCatAmt`), so nothing recomputes them and the drift is permanent in the workbook.
    # The classification totals reconcile in both months, which is what says the dates,
    # amounts and purchase types are sound and only the category landed wrong.
    ("October", "Expenses [Income]"): (
        "GBP 2,034.27 of 'Other' income added to the Expenses row: the four 31 October "
        "'Emma -' credits (food 190.66, excess 31.55, other 378.01, bills 1,434.05), named "
        "in the cell comment on C35. The same five credits in March are 'Other' in the "
        "workbook as well as the ledger, so the workbook disagrees with itself here."
    ),
    ("October", "Other [Income]"): "Counterpart of Expenses [Income]: the same GBP 2,034.27.",
    ("October", "Expenses [Spent]"): (
        "GBP 1,019.14 of 'Other' spend added to the Expenses row -- every 'Other' debit "
        "from 21 October to month end without exception, ten of them, from the 500.00 "
        "photographer deposit to the 31.05 haircut. Named in the cell comment on E35. "
        "Expenses' own GBP 1,307.00 IFOA conference charge is the remainder."
    ),
    ("October", "Other [Spent]"): "Counterpart of Expenses [Spent]: the same GBP 1,019.14.",
    ("March", "Expenses [Spent]"): (
        "GBP 249.95 of 'Other' spend added to the Expenses row: Debug rows 1824 (20.00), "
        "1827 (14.55), 1830 (34.90), 1832 (15.50) and 1833 (165.00), 3-5 March. Named one "
        "by one in the cell comment on E34. Expenses' own GBP 51.26 is the remainder."
    ),
    ("March", "Other [Spent]"): "Counterpart of Expenses [Spent]: the same GBP 249.95.",
    # The per-category view of corrections confirmed elsewhere. Nothing new -- the same
    # money, seen through the month tab's category totals rather than its balances.
    ("November", "Travel [Spent]"): (
        "Debug row 1221, 9 Nov, TfL recorded as 6.00 where it was 7.40. The corrected 1.40."
    ),
    ("February", "Lottery [Spent]"): (
        "Debug row 1796's GBP 85.05, which the workbook posted to the March tab."
    ),
    ("March", "Lottery [Spent]"): "The other end of row 1796: the same GBP 85.05, back out.",
    ("March", "Travel [Spent]"): (
        "Debug row 1929's 76p: the NY hotel charge at 499.85, not 500.61."
    ),
    # The Summary sheet deducts these by hand: D6 is `...-10000` and D7 `...-4130`, typed
    # into the formula. The money is really in the accounts -- 10,000 moved into Savings -
    # Halifax on 20 June and on to Savings - Wedding on 1 July, of which 4,130 was still
    # there at month end -- and section A reconciles both accounts in both months, so the
    # database is right and the workbook is netting off a pot it had no way to earmark.
    # By August the deduction stops and the two agree again.
    ("June", "Savings"): "Summary!D6 subtracts 10,000 in the formula itself.",
    ("July", "Savings"): "Summary!D7 subtracts 4,130 in the formula itself.",
}

# Running totals are cumulative, so every difference accepted in section B shows up again in
# the month's closing figure and in every month after it until something cancels it. These
# are those totals, and they are the sum of what is already above -- May's 28.60 is the four
# May daily differences added up, June adds row 373's 12.25, July rows 564 and 598, and
# February takes row 1796's 85.05 back out.
#
# Listed month by month rather than derived from the daily ones, so that a *new* difference
# in any month still fails the gate. Each appears twice in a run: once through Summary
# (check D) and once through the month tabs (check F), which is what the second loop builds.
_CUMULATIVE_25_26: dict[tuple[str, str], str] = {
    ("May", "Excess"): "28.60: -3.20 +7.95 +5.00 +18.85, the four accepted May days.",
    ("May", "Food"): "4.75: rows 190 and 194 swapped, -3.20 on the 1st and +7.95 on the 2nd.",
    ("June", "Bills"): "12.25: row 373, which the tab files as Excess.",
    ("June", "Excess"): "40.85: May's 28.60 plus row 373's 12.25 coming out of Excess.",
    ("July", "Bills"): "19.50: row 564, which the tab files as Excess.",
    ("July", "Excess"): "65.71: 40.85 plus row 564's 19.50 and row 598's 5.36.",
    ("August", "Excess"): "65.71, carried forward unchanged.",
    ("September", "Excess"): "65.71, carried forward unchanged.",
    ("October", "Excess"): "65.71, carried forward unchanged.",
    ("November", "Excess"): "65.71, carried forward unchanged.",
    ("December", "Excess"): "65.71, carried forward unchanged.",
    ("January", "Excess"): "65.71, carried forward unchanged.",
    ("February", "Excess"): "-19.34: 65.71 less row 1796's 85.05, dated February here.",
    ("March", "Excess"): "-19.34, carried forward unchanged.",
    ("March", "Savings"): "-0.76: row 1929's NY hotel charge at 499.85, not 500.61.",
}

ACCEPTED_25_26.update(_CUMULATIVE_25_26)
ACCEPTED_25_26.update(
    {(month, f"Running {key}"): note for (month, key), note in _CUMULATIVE_25_26.items()}
)

ACCEPTED_BY_YEAR = {2025: ACCEPTED_25_26, 2026: ACCEPTED_26_27}

# Differences that are understood but not yet decided. Listed so the run stays usable as a
# regression gate -- a *new* difference still fails -- while staying visible every time.
PENDING: dict[tuple[str, str], str] = {}


@dataclass
class Diff:
    month: str
    key: str
    workbook: Decimal
    computed: Decimal

    @property
    def delta(self) -> Decimal:
        return self.computed - self.workbook


def _fmt(value: Decimal) -> str:
    return f"{value:>12,.2f}"


# ------------------------------------------------------------------ A: account balances


def check_balances(
    session, values, formulas, ref: xr.RefData, verbose: bool
) -> tuple[list[Diff], dict[str, tuple[str, ...]]]:
    """Returns (differences, columns that matched no account).

    The second is separate because it is not a difference, it is a *gap*: nothing was
    compared at all. Left as a difference it would report an account's whole balance as
    wrong in every month, which is both alarming and uninformative; left out entirely the
    run would go green with an account never checked, which is worse.
    """
    accounts = {a.id: a for a in session.scalars(select(Account))}
    types = {a.name: a.type for a in accounts.values()}

    diffs: list[Diff] = []
    print("\nA. Account balances  (opening + postings  vs  month-tab 'End')")
    print("   " + "-" * 74)

    # Carried month to month rather than read per month. The stored opening_balance rows
    # were copied from the workbook's row 60, which is a formula -- the previous month's End
    # -- so as values they describe the day the migration ran and nothing since. Rolling
    # forward here is the workbook's own rule, computed independently of repo so this stays
    # a cross-check rather than a restatement of the query layer.
    running: dict[str, Decimal] = {}
    seeded: set[str] = set()
    # Month-tab columns that name no account at all. Reported on their own rather than as
    # balance differences: an account renamed in the dashboard would otherwise read as one
    # holding nothing, which is a difference of its whole balance in every month and says
    # nothing about the cause. Tembo, renamed to 'Savings - Tembo', cost an afternoon.
    unmatched: dict[str, tuple[str, ...]] = {}

    for month in xr.FISCAL_MONTHS:
        period = ref.period_for(month)
        layout = xr.month_layout(values, formulas, month, period)
        closing = xr.read_closing_balances(values, layout)

        stored = {
            accounts[ob.account_id].name: ob.amount
            for ob in session.scalars(
                select(OpeningBalance).where(OpeningBalance.period == period)
            )
        }
        # An account is seeded from its stated opening the first month it has one. For one
        # opened part-way through the year that is its real starting balance, with nothing
        # before it to roll forward from.
        for name, amount in stored.items():
            if name not in seeded:
                running[name] = amount
                seeded.add(name)
        opening = running

        movement: dict[str, Decimal] = defaultdict(Decimal)
        txns = session.scalars(
            select(Txn).where(Txn.period == period, Txn.deleted_at.is_(None))
        )
        n_txn = 0
        for t in txns:
            n_txn += 1
            to_name = accounts[t.account_to_id].name if t.account_to_id else None
            for posting in postings_for(
                t.type, accounts[t.account_from_id].name, to_name, t.amount
            ):
                movement[posting.account] += posting.signed(types[posting.account])

        month_diffs = []
        for block in layout.blocks:
            name = block.name
            # One column can stand for several accounts: a workbook that predates a split
            # shows the cards added together, which is what its 'End' figure is.
            parts = db_accounts_for_block(name)
            if not any(p in types for p in parts):
                # The alias named nothing. If the workbook's own name still matches an
                # account, use it -- an alias records what a column used to be called, and a
                # database written before the rename is not wrong.
                if name in types:
                    parts = (name,)
                else:
                    unmatched.setdefault(name, parts)
                    continue
            computed = sum(
                (opening.get(p, Decimal("0")) + movement.get(p, Decimal("0")) for p in parts),
                Decimal("0"),
            )
            expected = closing.get(name, Decimal("0"))
            if abs(computed - expected) > TOLERANCE:
                month_diffs.append(Diff(month, name, expected, computed))
            # This month's close is next month's open, carried per account rather than per
            # column so a split stays split. Carried whether or not it matched, so one bad
            # month shows as one difference rather than repeating to March.
            for part in parts:
                running[part] = opening.get(part, Decimal("0")) + movement.get(
                    part, Decimal("0")
                )

        status = "OK" if not month_diffs else f"{len(month_diffs)} MISMATCH"
        print(
            f"   {month:<10} {len(layout.blocks):>3} accounts  {n_txn:>4} txns   {status}"
        )
        for d in month_diffs:
            print(
                f"        {d.key:<28} workbook{_fmt(d.workbook)}"
                f"  computed{_fmt(d.computed)}  delta{_fmt(d.delta)}"
            )
        diffs += month_diffs

    return diffs, unmatched


# ------------------------------------------------------- B: daily classification totals


def check_classification_totals(
    session, values, ref: xr.RefData, verbose: bool
) -> list[Diff]:
    """Workbook rule, from the DB4 formula:

        (SUMIFS(debits...) - SUMIFS(credits...) + ...per account...) * DB$38

    where row 38 is INDEX(xlSelJ, MATCH(classification, xlSelH)) -- the direction from
    Selections!J. So: direction * (debits - credits). Only Excess carries direction -1,
    which is why it is the classification that shows up inverted if this is missed.

    Non-transfer transactions only: New_entry writes the classification reference only when
    the type is not 'Transfer', so transfers are excluded entirely rather than netting out.
    """
    classifications = list(session.scalars(select(Classification)))
    directions = {c.id: Decimal(c.direction) for c in classifications}

    diffs: list[Diff] = []
    print("\nB. Daily classification totals  (debits - credits  vs  month-tab totals)")
    print("   " + "-" * 74)

    for month in xr.FISCAL_MONTHS:
        period = ref.period_for(month)

        computed: dict[tuple[str, dt.date], Decimal] = defaultdict(Decimal)
        for t in session.scalars(
            select(Txn).where(
                Txn.period == period,
                Txn.deleted_at.is_(None),
                Txn.type != "Transfer",
                Txn.classification_id.is_not(None),
            )
        ):
            name = next(c.name for c in classifications if c.id == t.classification_id)
            sign = Decimal("1") if t.type == "Debit" else Decimal("-1")
            computed[(name, t.txn_date)] += directions[t.classification_id] * sign * t.amount

        month_diffs = []
        checked = 0
        for cls in classifications:
            try:
                workbook_totals = xr.read_daily_class_totals(values, month, cls.name)
            except KeyError:
                continue  # 'Other' has no per-day column in the workbook
            for day, expected in workbook_totals.items():
                checked += 1
                got = computed.get((cls.name, day), Decimal("0"))
                if abs(got - expected) > TOLERANCE:
                    month_diffs.append(Diff(month, f"{cls.name} {day}", expected, got))

        status = "OK" if not month_diffs else f"{len(month_diffs)} MISMATCH"
        print(f"   {month:<10} {checked:>4} cells   {status}")
        for d in month_diffs[: (None if verbose else 8)]:
            print(
                f"        {d.key:<28} workbook{_fmt(d.workbook)}"
                f"  computed{_fmt(d.computed)}  delta{_fmt(d.delta)}"
            )
        if not verbose and len(month_diffs) > 8:
            print(f"        ... {len(month_diffs) - 8} more (use -v)")
        diffs += month_diffs

    return diffs


# ------------------------------------------------- C: per-category spend (month tab col E)


def check_category_spend(session, values, ref: xr.RefData, verbose: bool) -> list[Diff]:
    """Month-tab columns C ('Income') and E ('Total Spent') -- what the Month page shows
    against budget.

    Note these are macro-maintained accumulators, not formulas: New_entry does
    `ActiveCell = dblAmt + dblCatAmt`. They can therefore drift from the ledger in a way a
    formula could not, which is exactly what makes them worth checking.
    """
    postings = repo.load_postings(session)

    diffs: list[Diff] = []
    print("\nC. Per-category income and spend  (month-tab 'Income' / 'Total Spent')")
    print("   " + "-" * 74)

    for month in xr.FISCAL_MONTHS:
        period = ref.period_for(month)
        sheet, col, min_row, _, max_row = xr.resolve(values, f"xlCategories{month}")
        ws = values[sheet]

        actuals = repo.category_actuals(postings, period)
        computed = actuals.set_index("category") if not actuals.empty else actuals

        month_diffs, checked = [], 0
        for r in range(min_row, max_row + 1):
            name = ws.cell(r, col).value
            if not name:
                continue
            name = str(name).strip()
            for offset, column, label in ((1, "income", "Income"), (3, "spent", "Spent")):
                raw = ws.cell(r, col + offset).value
                if raw in (None, "-", ""):
                    continue  # '-' marks a column that does not apply to this category
                checked += 1
                expected = Decimal(str(raw))
                got = Decimal("0")
                if not computed.empty and name in computed.index:
                    got = computed[column].get(name, Decimal("0"))
                if abs(got - expected) > TOLERANCE:
                    month_diffs.append(Diff(month, f"{name} [{label}]", expected, got))

        status = "OK" if not month_diffs else f"{len(month_diffs)} MISMATCH"
        print(f"   {month:<10} {checked:>4} cells   {status}")
        for d in month_diffs[: (None if verbose else 8)]:
            print(
                f"        {d.key:<32} workbook{_fmt(d.workbook)}"
                f"  computed{_fmt(d.computed)}  delta{_fmt(d.delta)}"
            )
        diffs += month_diffs

    return diffs


# ------------------------------------------- D: classification by month (Summary Q19:Y31)


def _running_closes(session, ref: xr.RefData):
    """Month-end running totals from the rollover engine (DESIGN.md 6a)."""
    reference = repo.load_reference(session)
    retention = Decimal(str(reference["settings"].get("excess_retention", 1)))
    return repo.running_by_period(
        repo.load_postings(session),
        repo.load_projections(session),
        repo.load_allowances(session),
        reference["classifications"],
        repo.fiscal_periods(ref.tax_year),
        retention,
        openings=repo.load_class_openings(session),
    )


def check_summary_matrix(session, values, ref: xr.RefData, verbose: bool) -> list[Diff]:
    """Summary!R20:Y31, which reads `xlCloseValue` -- the cumulative closing position, not
    monthly actuals. Prior months rolled forward per each classification's rollover rule,
    plus projections for dates after today.

    Skipped through Phases 1-3 because reproducing it needs the rollover engine and the
    projection table. Both now exist, so it is a live check again.
    """
    ws = values["Summary"]
    _, closes = _running_closes(session, ref)
    lookup = closes.set_index(["period", "classification"])["closing"]

    month_col, headers, rows = xr.summary_matrix(values, ref)

    diffs: list[Diff] = []
    print(
        f"\nD. Summary classification matrix  (Summary!"
        f"{get_column_letter(month_col)}{rows.start}:"
        f"{get_column_letter(max(headers))}{rows.stop - 1}, cumulative + projected)"
    )
    print("   " + "-" * 74)

    checked = 0
    for row in rows:
        month = ws.cell(row, month_col).value
        if not month:
            continue
        month = str(month).strip()
        period = ref.period_for(month)
        for col, name in headers.items():
            raw = ws.cell(row, col).value
            if raw is None or (period, name) not in lookup.index:
                continue  # 'Spend' is a derived total, not a classification
            checked += 1
            expected = Decimal(str(raw))
            got = Decimal(str(lookup.loc[(period, name)]))
            if abs(got - expected) > TOLERANCE:
                diffs.append(Diff(month, name, expected, got))

    status = "OK" if not diffs else f"{len(diffs)} MISMATCH"
    print(f"   {checked:>4} cells   {status}")
    for d in diffs[: (None if verbose else 12)]:
        print(
            f"        {d.month:<10} {d.key:<18} workbook{_fmt(d.workbook)}"
            f"  computed{_fmt(d.computed)}  delta{_fmt(d.delta)}"
        )
    return diffs


# ------------------------------------ E: savings and investment position (Summary D and K)


def check_savings_position(session, values, ref: xr.RefData, verbose: bool) -> list[Diff]:
    """Summary columns D and K -- the headline figures on the Overview page.

    These are driven entirely by the is_savings / is_investment flags on each account, so
    this check catches a mis-imported flag as well as a mis-computed balance.
    """
    ws = values["Summary"]
    postings = repo.load_postings(session)
    reference = repo.load_reference(session)
    openings = repo.load_opening_balances(session)

    diffs: list[Diff] = []
    print("\nE. Savings and investment position  (Summary!D and K)")
    print("   " + "-" * 74)

    today = dt.date.today()
    checked = skipped = 0
    for row in range(4, 16):  # row 3 is the 1 April opening; 4..15 are month ends
        month = ws.cell(row, 3).value  # C: Month
        if not month:
            continue
        month = str(month).strip()

        # Both formulas are =IF(TODAY()>=<month end>, SUM(...), 0), so a month that has not
        # finished reports zero by design rather than by omission. Comparing against those
        # zeroes would mean asserting the workbook's own placeholder.
        month_end = ws.cell(row, 2).value  # B: month-end date
        month_end = month_end.date() if isinstance(month_end, dt.datetime) else month_end
        if month_end and today < month_end:
            skipped += 1
            continue

        period = ref.period_for(month)
        balances = repo.account_balances(postings, openings, period, reference["accounts"])
        position = repo.savings_position(balances)

        for col, key, label in ((4, "savings", "Savings"), (11, "investments", "Investments")):
            raw = ws.cell(row, col).value
            if raw is None:
                continue
            checked += 1
            expected = Decimal(str(raw))
            got = position[key]
            if abs(got - expected) > TOLERANCE:
                diffs.append(Diff(month, label, expected, got))

    status = "OK" if not diffs else f"{len(diffs)} MISMATCH"
    print(f"   {checked:>4} cells   {status}   ({skipped} future month-ends skipped)")
    for d in diffs[: (None if verbose else 12)]:
        print(
            f"        {d.month:<10} {d.key:<14} workbook{_fmt(d.workbook)}"
            f"  computed{_fmt(d.computed)}  delta{_fmt(d.delta)}"
        )
    return diffs


# --------------------------------------- F: running classification totals (month tab CT:DA)


def check_running_totals(session, values, ref: xr.RefData, verbose: bool) -> list[Diff]:
    """Month-end value of each 'Running <Classification>' column.

    This is the rollover engine end to end: opening balance, carry-forward rule, retention,
    daily allowance, actuals to date and projections beyond. Check D reads the same figures
    through Summary, but this compares against the month tabs directly, so a discrepancy
    points at the month rather than at the summary that aggregates it.
    """
    _, closes = _running_closes(session, ref)
    lookup = closes.set_index(["period", "classification"])["closing"]

    diffs: list[Diff] = []
    print("\nF. Running classification totals  (month-tab 'Running ...' columns)")
    print("   " + "-" * 74)

    for month in xr.FISCAL_MONTHS:
        period = ref.period_for(month)
        sheet, min_col, _, max_col, _ = xr.resolve(values, f"xlClass{month}")
        ws = values[sheet]
        _, _, _, _, last_row = xr.resolve(values, f"xlCatDates{month}")

        month_diffs, checked = [], 0
        for col in range(min_col, max_col + 1):
            header = ws.cell(1, col).value
            if not header or not str(header).startswith("Running "):
                continue
            name = str(header)[len("Running "):]
            if (period, name) not in lookup.index:
                continue
            checked += 1
            expected = Decimal(str(ws.cell(last_row, col).value or 0))
            got = Decimal(str(lookup.loc[(period, name)]))
            if abs(got - expected) > TOLERANCE:
                month_diffs.append(Diff(month, f"Running {name}", expected, got))

        status = "OK" if not month_diffs else f"{len(month_diffs)} MISMATCH"
        print(f"   {month:<10} {checked:>4} columns   {status}")
        for d in month_diffs[: (None if verbose else 8)]:
            print(
                f"        {d.key:<24} workbook{_fmt(d.workbook)}"
                f"  computed{_fmt(d.computed)}  delta{_fmt(d.delta)}"
            )
        diffs += month_diffs

    return diffs


# ------------------------------------------------------------------------------- driver


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Reconcile the database against the workbook.")
    parser.add_argument("--workbook", type=Path, default=config.WORKBOOK_PATH)
    parser.add_argument("--db", type=Path, default=config.DB_PATH)
    parser.add_argument("-v", "--verbose", action="store_true", help="list every difference")
    args = parser.parse_args(argv)

    print(f"Workbook {args.workbook}")
    print(f"Database {args.db}")

    values, formulas = xr.load(args.workbook)
    ref = xr.read_reference(values)

    engine = make_engine(args.db)
    factory = make_session_factory(engine)

    with factory() as session:
        balance_diffs, unmatched = check_balances(session, values, formulas, ref, args.verbose)
        class_diffs = check_classification_totals(session, values, ref, args.verbose)
        category_diffs = check_category_spend(session, values, ref, args.verbose)
        summary_diffs = check_summary_matrix(session, values, ref, args.verbose)
        position_diffs = check_savings_position(session, values, ref, args.verbose)
        running_diffs = check_running_totals(session, values, ref, args.verbose)

    all_diffs = (
        balance_diffs + class_diffs + category_diffs + summary_diffs
        + position_diffs + running_diffs
    )
    # This workbook's own list, never another year's. See ACCEPTED_26_27.
    known = ACCEPTED_BY_YEAR.get(ref.tax_year, {})
    accepted = [d for d in all_diffs if (d.month, d.key) in known]
    pending = [d for d in all_diffs if (d.month, d.key) in PENDING]
    unexplained = [
        d for d in all_diffs if (d.month, d.key) not in known and (d.month, d.key) not in PENDING
    ]

    print("\n" + "=" * 78)
    if unmatched:
        # Its own failure, and ahead of everything else: a column nothing was compared for
        # is a hole in the gate, not a figure that disagrees. Reporting it as a difference
        # would say an account's whole balance was wrong in every month; reporting nothing
        # would let the run go green with an account never checked at all.
        print(f"{len(unmatched)} month-tab column(s) matched no account, so NOTHING WAS")
        print("COMPARED for them. Almost always a rename -- an account renamed in the")
        print("dashboard still has its old name in the workbook. Add it to ACCOUNT_ALIASES")
        print("in budget/backfill_year.py, or rename the account back.\n")
        for name, tried in sorted(unmatched.items()):
            looked = ", ".join(repr(p) for p in tried)
            print(f"  column {name!r}: no account called {looked}")
        print()
        return 1

    if accepted:
        print(f"{len(accepted)} accepted difference(s) -- database deliberately differs:")
        for d in accepted:
            print(f"  {d.month} {d.key}  delta{_fmt(d.delta)}")
            print(f"    {known[(d.month, d.key)]}")
        print()
    if pending:
        print(f"{len(pending)} PENDING difference(s) -- understood, awaiting a decision:")
        for d in pending:
            print(f"  {d.month} {d.key}  delta{_fmt(d.delta)}")
            print(f"    {PENDING[(d.month, d.key)]}")
        print()

    if not unexplained:
        print("RECONCILED: the ledger reproduces the workbook, bar the accepted differences.")
        return 0

    buckets = [
        ("balance", balance_diffs),
        ("daily total", class_diffs),
        ("category spend", category_diffs),
        ("summary matrix", summary_diffs),
        ("savings position", position_diffs),
        ("running totals", running_diffs),
    ]
    parts = [
        f"{len([d for d in unexplained if d in group])} {label}"
        for label, group in buckets
        if any(d in group for d in unexplained)
    ]
    print(f"{len(unexplained)} unexplained: " + ", ".join(parts))
    print(f"largest absolute difference: {max(abs(d.delta) for d in unexplained):,.2f}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
