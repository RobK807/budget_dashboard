"""Seeds the Phase 5 tables from the workbook.

Phase 5 turned several things that were constants inside formulas into data: the annual
salary and the bonus behind 'Gross pay (expected)', the cycling fares, the credit-card
statement figures, and the savings and account targets. This reads them out of the workbook
one last time so an existing database starts from the same numbers rather than from blanks.

Like import_phase4 it touches only its own tables and bumps the revision, so it syncs like
any other edit rather than forcing a rebuild.

Run with:  python -m budget.import_phase5 [--db PATH] [--workbook PATH]
"""

from __future__ import annotations

import argparse
import datetime as dt
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from budget import config, service
from budget.db import create_all, make_engine, make_session_factory
from budget.models import (
    Account,
    AccountTarget,
    Bonus,
    Card,
    CardStatement,
    CyclingRate,
    Payslip,
    SalaryAssumption,
    SalaryProfile,
    SavingsTarget,
    Setting,
)

# The three savings pots the Summary tab's 'Less SC & Wed' column subtracted by hand. Tembo
# joined in June and the header was never updated -- the arithmetic did include it, so the
# label was wrong rather than the figure. Verified against all four months of the workbook.
EARMARKED = ("Savings - Service Charge", "Savings - Wedding", "Tembo")

CYCLING_START = dt.date(2026, 4, 1)

# Salary tracker H17:I28. Bills and other costs are derived from the month's budget; these
# four are the inputs, taken from the workbook's own figures.
SPENDING_DEFAULTS = {
    "spend_rent": "0",
    "spend_savings": "1000",
    "spend_food": "500",
    "spend_essentials": "50",
}


def _dec(value) -> Decimal | None:
    if value is None or value == "":
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _date(value) -> dt.date | None:
    if value is None:
        return None
    return value.date() if isinstance(value, dt.datetime) else value


def read_card_months(values) -> tuple[dict[str, dict[str, Decimal]], dict[str, dict]]:
    """Month-tab rows 42-50: the statement figures and the two day-of-month settings.

    The block sits at a different row in each month tab, so it is found by its label rather
    than by a fixed offset -- the same approach read_daily_allowance takes.
    """
    statements: dict[str, dict[str, Decimal]] = {}
    days: dict[str, dict] = {}

    for month in (
        "April", "May", "June", "July", "August", "September",
        "October", "November", "December", "January", "February", "March",
    ):
        if month not in values.sheetnames:
            continue
        ws = values[month]
        label_row = None
        for row in range(30, 140):
            if ws.cell(row, 2).value == "Credit card bill BoM":
                label_row = row
                break
        if label_row is None:
            continue

        header_row = label_row - 2
        for column in range(3, 9):
            name = ws.cell(header_row, column).value
            if not name:
                continue
            name = str(name).strip()
            eom = _dec(ws.cell(label_row + 2, column).value)
            if eom:
                statements.setdefault(month, {})[name] = eom
            statement_day = ws.cell(label_row + 4, column).value
            payment_day = ws.cell(label_row + 6, column).value
            if name not in days and statement_day:
                days[name] = {
                    "statement_day": int(statement_day),
                    "payment_day": int(payment_day) if payment_day else None,
                }
    return statements, days


def read_summary_targets(values) -> tuple[dict[str, Decimal], dict[int, dict]]:
    """Summary C23:C26 (per-account) and G3:G15 / M3:M15 (savings and investments)."""
    ws = values["Summary"]

    account_targets: dict[str, Decimal] = {}
    for row in range(23, 30):
        name = ws.cell(row, 2).value
        if not name:
            continue
        amount = _dec(ws.cell(row, 3).value)
        if amount is not None:
            account_targets[str(name).strip()] = amount

    savings_targets: dict[int, dict] = {}
    for row in range(3, 16):
        when = _date(ws.cell(row, 2).value)
        if when is None:
            continue
        savings_targets[row] = {
            "date": when,
            "savings": _dec(ws.cell(row, 7).value),      # G
            "investments": _dec(ws.cell(row, 13).value),  # M
        }
    return account_targets, savings_targets


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Seed the Phase 5 tables.")
    parser.add_argument("--workbook", type=Path, default=config.WORKBOOK_PATH)
    parser.add_argument("--db", type=Path, default=config.DB_PATH)
    args = parser.parse_args(argv)

    print(f"Reading  {args.workbook}")
    values = load_workbook(args.workbook, data_only=True)

    engine = make_engine(args.db)
    # create_all migrates first and then creates: the ALTERs land on the existing tables
    # before the new ones are built.
    for note in create_all(engine):
        print(f"  schema: {note}")
    factory = make_session_factory(engine)

    counts: dict[str, int] = {}

    with factory() as session, session.begin():
        accounts = {a.name: a for a in session.scalars(select(Account))}

        # --- earmarked savings pots -------------------------------------------------
        marked = 0
        for name in EARMARKED:
            account = accounts.get(name)
            if account is not None and not account.exclude_from_savings:
                account.exclude_from_savings = True
                marked += 1
        counts["earmarked"] = marked

        # --- salary profile and bonuses ---------------------------------------------
        session.execute(delete(SalaryProfile))
        session.execute(delete(Bonus))

        payslips = sorted(session.scalars(select(Payslip)), key=lambda p: p.period)
        previous_salary = None
        profiles = bonuses = 0
        for payslip in payslips:
            if payslip.salary is None:
                continue
            year, month = (int(p) for p in payslip.period.split("-"))
            if previous_salary is None or payslip.salary != previous_salary:
                session.add(
                    SalaryProfile(
                        effective_from=dt.date(year, month, 1),
                        annual_salary=payslip.salary,
                        note="From the workbook's Salary tracker column O",
                    )
                )
                previous_salary = payslip.salary
                profiles += 1

            # Whatever the stated expected gross carries above salary/12 is the bonus --
            # May's cell was `=ROUND(O5/12,2)+29028.48`, the figure welded into the formula.
            if payslip.expected_gross is not None:
                baseline = (payslip.salary / 12).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                extra = payslip.expected_gross - baseline
                if extra:
                    session.add(
                        Bonus(period=payslip.period, amount=extra, note="Bonus")
                    )
                    bonuses += 1
        counts["salary profiles"] = profiles
        counts["bonuses"] = bonuses

        # --- the two band inputs behind the basic band ------------------------------
        # Phase 4 stored only D36 ('Adjusted bands' -> Basic rate), which is D28 - D22.
        # The editable table needs the inputs, not just the figure derived from them.
        ws = values["Salary tracker"]
        tax_year = int(session.get(Setting, "tax_year").value)
        added = 0
        for key, row in (("personal_allowance", 22), ("basic_rate_threshold", 28)):
            existing = session.get(
                SalaryAssumption, (tax_year, key, dt.date(tax_year, 4, 1))
            )
            value = _dec(ws.cell(row, 4).value)
            if existing is None and value is not None:
                session.add(
                    SalaryAssumption(
                        tax_year=tax_year,
                        key=key,
                        effective_from=dt.date(tax_year, 4, 1),
                        value=value,
                    )
                )
                added += 1
        counts["band inputs"] = added

        # --- cycling rates ----------------------------------------------------------
        session.execute(delete(CyclingRate))
        rates = 0
        for kind in ("commute", "band", "gym"):
            setting = session.get(Setting, f"cycling_rate_{kind}")
            if setting is None:
                continue
            session.add(
                CyclingRate(
                    kind=kind,
                    effective_from=CYCLING_START,
                    amount=Decimal(setting.value),
                )
            )
            rates += 1
        counts["cycling rates"] = rates

        # --- card statements and the two day-of-month settings ----------------------
        session.execute(delete(CardStatement))
        statements, days = read_card_months(values)

        month_periods = {}
        for payslip in payslips:
            year, month = (int(p) for p in payslip.period.split("-"))
            month_periods[dt.date(year, month, 1).strftime("%B")] = payslip.period

        written = 0
        for month, per_card in statements.items():
            period = month_periods.get(month)
            if period is None:
                continue
            for name, amount in per_card.items():
                account = accounts.get(name)
                if account is None:
                    continue
                session.add(
                    CardStatement(
                        period=period, account_id=account.id, bill_eom=amount
                    )
                )
                written += 1
        counts["card statements"] = written

        dated = 0
        for name, setting in days.items():
            account = accounts.get(name)
            if account is None:
                continue
            account.statement_day = setting["statement_day"]
            account.payment_day = setting["payment_day"]
            dated += 1
        counts["card payment days"] = dated

        # --- targets ----------------------------------------------------------------
        session.execute(delete(AccountTarget))
        session.execute(delete(SavingsTarget))
        account_targets, savings_targets = read_summary_targets(values)

        periods = sorted(month_periods.values())
        current = dt.date.today().strftime("%Y-%m")

        # The workbook held one set of account targets, for whichever month it happened to
        # be showing, and they were meant to apply month after month. Copied across every
        # month up to the current one so the table is populated whichever is selected;
        # from here each month keeps its own and can diverge.
        written = 0
        applies_to = [p for p in periods if p <= current] or [current]
        for name, amount in account_targets.items():
            account = accounts.get(name)
            if account is None:
                continue
            for period in applies_to:
                session.add(
                    AccountTarget(period=period, account_id=account.id, amount=amount)
                )
                written += 1
        counts["account targets"] = written

        written = 0
        for row in savings_targets.values():
            period = row["date"].strftime("%Y-%m")
            if periods and period < periods[0]:
                continue
            if row["savings"] is None and row["investments"] is None:
                continue
            existing = session.get(SavingsTarget, period)
            if existing is None:
                session.add(
                    SavingsTarget(
                        period=period,
                        savings=row["savings"],
                        investments=row["investments"],
                    )
                )
                written += 1
        counts["savings targets"] = written

        # --- credit limits ----------------------------------------------------------
        # Not imported: the workbook's 'Total available' column is a derived figure that
        # goes negative (`=9000-C4/1.033`), not a credit limit. Left blank to be entered.
        counts["cards without a credit limit"] = len(
            [c for c in session.scalars(select(Card)) if c.credit_limit is None]
        )

        # --- spending calculation inputs --------------------------------------------
        for key, value in SPENDING_DEFAULTS.items():
            if session.get(Setting, key) is None:
                session.add(Setting(key=key, value=value))

        revision = service.bump_revision(session)

    print(f"Wrote    {args.db}")
    for label, count in counts.items():
        print(f"  {label + ':':<28} {count}")
    print(f"  revision now: {revision} — push from the Sync page")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
