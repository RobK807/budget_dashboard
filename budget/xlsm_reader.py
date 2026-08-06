"""Reads the legacy workbook.

Layout is discovered from the workbook's own defined names and header rows rather than
from hardcoded coordinates. That matters because the month tabs are not uniform: April
and May carry extra inserted rows and fewer accounts than August onwards, so
xlCatDatesApril is CG4:CG33 where xlCatDatesAugust is CS4:CS34. Anchoring on the names
the macros themselves use makes the reader indifferent to that.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.workbook.workbook import Workbook

# Fiscal year runs April -> March.
FISCAL_MONTHS = [
    "April", "May", "June", "July", "August", "September",
    "October", "November", "December", "January", "February", "March",
]

FIRST_ACCOUNT_COL = 8  # column H on every month tab
COLS_PER_ACCOUNT = 4  # Account | Credit | Debit | Date


@dataclass
class AccountRef:
    name: str
    short_code: str
    offset: int
    is_savings: bool
    savings_limit: Decimal
    is_investment: bool
    investment_limit: Decimal
    is_isa: bool
    display_order: int
    first_month: str | None  # first fiscal month the account appears in


@dataclass
class CategoryRef:
    name: str
    grouping: str
    spend_type: str
    display_order: int


@dataclass
class ClassificationRef:
    name: str
    legacy_ref: int
    direction: int
    display_order: int
    rollover: str = "none"


@dataclass
class RefData:
    accounts: list[AccountRef]
    categories: list[CategoryRef]
    classifications: list[ClassificationRef]
    months: dict[str, int]  # name -> month number
    settings: dict[str, str]

    @property
    def tax_year(self) -> int:
        return int(self.settings["tax_year"])

    def period_for(self, month_name: str) -> str:
        """'August' -> '2026-08'. Jan-Mar belong to the following calendar year."""
        num = self.months[month_name]
        year = self.tax_year if num >= 4 else self.tax_year + 1
        return f"{year:04d}-{num:02d}"


@dataclass
class LedgerRow:
    source_row: int
    txn_date: dt.date
    identifier: str
    month: str
    account_from: str
    account_to: str | None
    type: str
    category: str | None
    amount: Decimal
    category_comment: str | None
    comment: str | None
    classification: str | None
    created_at: dt.datetime | None
    removed: bool
    removed_reason: str | None = None


@dataclass
class AccountBlock:
    name: str
    base_col: int
    type: str  # 'bank' | 'credit_card'


@dataclass
class MonthLayout:
    name: str
    period: str
    start_row: int
    end_row: int
    blocks: list[AccountBlock] = field(default_factory=list)

    def block(self, account_name: str) -> AccountBlock | None:
        for b in self.blocks:
            if b.name == account_name:
                return b
        return None


# ------------------------------------------------------------------ workbook plumbing


def load(path: Path) -> tuple[Workbook, Workbook]:
    """Return (values, formulas) views of the workbook."""
    values = openpyxl.load_workbook(path, data_only=True, read_only=False)
    formulas = openpyxl.load_workbook(path, data_only=False, read_only=False)
    return values, formulas


def resolve(wb: Workbook, name: str) -> tuple[str, int, int, int, int]:
    """Defined name -> (sheet, min_col, min_row, max_col, max_row)."""
    sheet, coord = next(iter(wb.defined_names[name].destinations))
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    return sheet, min_col, min_row, max_col, max_row


def _clean(value):
    """The workbook uses 0 as a null for optional text fields."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if value == 0:
        return None
    return value


def _dec(value) -> Decimal:
    if value is None or value == "-" or value == "":
        return Decimal("0")
    return Decimal(str(value))


# ------------------------------------------------------------------- reference tables


def read_reference(values: Workbook) -> RefData:
    sel = values["Selections"]
    ctl = values["Control"]

    months: dict[str, int] = {}
    for r in range(4, 16):
        nm = sel.cell(r, 1).value  # A: Month
        if nm:
            months[str(nm)] = int(sel.cell(r, 16).value)  # P: MonthNum

    # Per-month account offsets live in Z:AK (columns 26..37) in fiscal month order. A zero
    # means the account did not exist that month, which is what valid_from is derived from.
    accounts: list[AccountRef] = []
    for i, r in enumerate(range(4, 26)):
        name = _clean(sel.cell(r, 4).value)  # D: Account
        if not name:
            continue
        first_month = None
        for m_idx, month in enumerate(FISCAL_MONTHS):
            if (sel.cell(r, 26 + m_idx).value or 0) != 0:
                first_month = month
                break
        accounts.append(
            AccountRef(
                name=str(name),
                short_code=str(sel.cell(r, 15).value),  # O: Identifiers
                offset=int(sel.cell(r, 5).value or 0),  # E: Offset
                is_savings=bool(sel.cell(r, 18).value),  # R
                savings_limit=_dec(sel.cell(r, 19).value),  # S
                is_investment=bool(sel.cell(r, 20).value),  # T
                investment_limit=_dec(sel.cell(r, 21).value),  # U
                is_isa=bool(sel.cell(r, 22).value),  # V
                display_order=i,
                first_month=first_month,
            )
        )

    categories: list[CategoryRef] = []
    for i, r in enumerate(range(4, 40)):
        name = _clean(sel.cell(r, 7).value)  # G: Category
        if not name:
            continue
        categories.append(
            CategoryRef(
                name=str(name),
                grouping=str(sel.cell(r, 39).value),  # AM: CategoryGroupings
                spend_type=str(sel.cell(r, 41).value),  # AO: SpendingType
                display_order=i,
            )
        )

    classifications: list[ClassificationRef] = []
    for i, r in enumerate(range(4, 12)):
        name = _clean(sel.cell(r, 8).value)  # H: Categories (classifications)
        if not name:
            continue
        classifications.append(
            ClassificationRef(
                name=str(name),
                legacy_ref=int(sel.cell(r, 9).value),  # I: Reference
                direction=int(sel.cell(r, 10).value),  # J: Direction
                display_order=i,
            )
        )

    settings = {
        "tax_year": str(int(ctl["I4"].value)),
        "user": str(ctl["I5"].value),
        "month_start_day": str(int(ctl["I6"].value)),
        "excess_retention": str(ctl["I7"].value),
    }

    return RefData(accounts, categories, classifications, months, settings)


# Behavioural, not literal. Selections!AL named the sign of a running total, but the
# formula behind 'Negative' never restricted to negative values -- MIN(0, ...) appears only
# in the older April/May/June form. From July it reads
#
#     prior_close * IF(prior_close > 0, xlExcessRetention, 1)
#
# which carries *everything*, with retention on a surplus. That is 'all' plus retention in
# this model, so importing it as 'debit' would reproduce the word and lose the behaviour.
# 'Positive' is MAX(0, prior_close): genuinely restricted, and for Excess (direction -1) a
# positive running total is a credit balance.
#
# See DESIGN.md 6a.
ROLLOVER_MAP = {"positive": "credit", "negative": "all", "all": "all", "none": "none"}


def read_rollovers(values: Workbook, month: str) -> dict[str, str]:
    """Rollover type per classification, held at row 36 of the running-total columns."""
    sheet, min_col, _, max_col, _ = resolve(values, f"xlClass{month}")
    ws = values[sheet]
    out: dict[str, str] = {}
    for col in range(min_col, max_col + 1):
        header = ws.cell(1, col).value
        rollover = str(ws.cell(36, col).value or "None").lower()
        if header and str(header).startswith("Running "):
            out[str(header)[len("Running "):]] = ROLLOVER_MAP.get(rollover, "none")
    return out


# -------------------------------------------------------------------------- the ledger


# The three known bad dates, keyed by source row so a changed workbook cannot silently
# misapply them (DESIGN.md 4, 'Date correction'). Deliberately enumerated rather than a
# blanket year coercion, which would corrupt genuine prior-year data on backfill.
DATE_FIXES: dict[int, tuple[dt.date, dt.date]] = {
    2: (dt.date(2019, 4, 1), dt.date(2026, 4, 1)),
    360: (dt.date(2029, 5, 29), dt.date(2026, 5, 29)),
    601: (dt.date(2024, 7, 11), dt.date(2026, 7, 11)),
}

# Where the ledger and the month tab disagree, the month tab wins -- it is the figure the
# balances were reconciled against and the one the user acts on. Confirmed case by case
# during Phase 0 reconciliation; see PHASE0_FINDINGS.md finding 3.
AMOUNT_FIXES: dict[int, tuple[Decimal, Decimal]] = {
    592: (Decimal("0.1"), Decimal("8.00")),  # 3 Jul, BA Amex, Travel
}

# Same idea for a miscategorised row. Debug row 533 is a GBP 75 birthday gift filed against
# 'Omaze' -- a GBP 15/month lottery subscription -- where the month tab and its cell comment
# both say 'Other'. See PHASE1_NOTES.md.
CATEGORY_FIXES: dict[int, tuple[str, str]] = {
    533: ("Omaze", "Other"),  # 25 Jun, HSBC, 'Emma birthday - M&D'
}

# Rows present in the ledger but never posted to any month tab. Imported so the audit trail
# survives, but soft-deleted so they carry no weight; see PHASE0_FINDINGS.md finding 1.
PHANTOM_ROWS: dict[int, str] = {
    2: "Template seed row: no matching entry in the April tab",
}


@dataclass(frozen=True)
class Corrections:
    """The known bad rows of one workbook, and the year its ledger belongs to.

    Every fix is keyed by source row and states the value it expects to find, so a workbook
    that has changed since the fix was confirmed raises rather than silently correcting the
    wrong row. That contract is per *workbook*: 25-26 and 26-27 are different files with
    different bad rows, and applying one file's fixes to the other would be exactly the
    silent corruption the row check exists to prevent.

    Backfilling is why this is a parameter rather than four module constants. The year check
    used to be `date.year != 2026`, which rejected every genuine 2025 date -- and would have
    rejected January 2027 too, once the current year runs that far.
    """

    tax_year: int
    dates: dict[int, tuple[dt.date, dt.date]] = field(default_factory=dict)
    amounts: dict[int, tuple[Decimal, Decimal]] = field(default_factory=dict)
    categories: dict[int, tuple[str, str]] = field(default_factory=dict)
    accounts: dict[int, tuple[str, str]] = field(default_factory=dict)
    phantoms: dict[int, str] = field(default_factory=dict)

    def covers(self, date: dt.date) -> bool:
        """April of the tax year through March of the next -- the fiscal year, not the
        calendar one."""
        return dt.date(self.tax_year, 4, 1) <= date <= dt.date(self.tax_year + 1, 3, 31)


CORRECTIONS_26_27 = Corrections(
    tax_year=2026,
    dates=DATE_FIXES,
    amounts=AMOUNT_FIXES,
    categories=CATEGORY_FIXES,
    phantoms=PHANTOM_ROWS,
)

# Budget 25-26.xlsm. Confirmed with the user on 6 August 2026:
#   row 2     the same template seed row as 26-27 -- imported for the audit trail and
#             soft-deleted, so it carries no weight
#   row 1169  GBP 15 Omaze on Amex, identifier 1101_AME_0 and month 'November', so only the
#             year was wrong
#   row 1929  GBP 500.61 NY hotel on 29 March; the real figure is 499.85, and the 76p is
#             exactly what stopped Platinum Amex's opening reconciling across the year end
#   row 326   an exact duplicate of row 324 -- the same HSBC -> First Direct transfer of
#             3,524.58 on the same day, entered twice. Kept and soft-deleted rather than
#             dropped, so the ledger still shows what the workbook contained.
#   row 372   6 June, 148.40 'Hyrox' -- on Mastercard, not Amex
#   row 589   9 July, a TfL debit left at zero; it was 10.50
#   row 1221  9 November, TfL recorded as 6.00; it was 7.40
#   row 1547  a TfL debit left at zero and dated a day early: 7.70 on 9 January
#   rows 1545, 1546  the two entries either side of it, a day early for the same reason:
#             21.00 'Theatre' and 1.50 'Blackwall tunnel' belong on 9 January. The January
#             tab totals 30.20 of Excess on the 9th, which is exactly these three added up.
# Row 1796 (27 February, 85.05 'Lottery') is *not* corrected: the ledger is right and the
# workbook put it on the March tab. It is an accepted difference in reconcile.py instead.
#
# The user counts the header as row 1, so their numbering runs one behind openpyxl's. These
# are openpyxl's.
CORRECTIONS_25_26 = Corrections(
    tax_year=2025,
    dates={
        2: (dt.date(2019, 4, 1), dt.date(2025, 4, 1)),
        1169: (dt.date(2022, 11, 1), dt.date(2025, 11, 1)),
        1545: (dt.date(2026, 1, 8), dt.date(2026, 1, 9)),
        1546: (dt.date(2026, 1, 8), dt.date(2026, 1, 9)),
        1547: (dt.date(2026, 1, 8), dt.date(2026, 1, 9)),
    },
    amounts={
        589: (Decimal("0"), Decimal("10.50")),
        1221: (Decimal("6"), Decimal("7.40")),
        1547: (Decimal("0"), Decimal("7.70")),
        1929: (Decimal("500.61"), Decimal("499.85")),
    },
    accounts={372: ("Amex", "Mastercard")},
    phantoms={
        2: "Template seed row: no matching entry in the April tab",
        326: "Duplicate of row 324: the same HSBC to First Direct transfer, entered twice",
    },
)


def read_ledger(
    values: Workbook, corrections: Corrections = CORRECTIONS_26_27
) -> tuple[list[LedgerRow], list[str]]:
    """Read the Debug sheet positionally.

    Columns B and C are read by position, not header: New_entry writes the identifier to B
    and the month to C, but the header labels say the opposite.
    """
    ws = values["Debug"]
    rows: list[LedgerRow] = []
    notes: list[str] = []
    applied_fixes = applied_amount_fixes = applied_phantoms = applied_category_fixes = 0
    applied_account_fixes = 0

    for r in range(2, ws.max_row + 1):
        raw_date = ws.cell(r, 1).value
        if raw_date is None:
            break

        date = raw_date.date() if isinstance(raw_date, dt.datetime) else raw_date

        if r in corrections.dates:
            expected, corrected = corrections.dates[r]
            if date != expected:
                raise ValueError(
                    f"Debug row {r}: expected known-bad date {expected}, found {date}. "
                    "The workbook has changed -- re-verify the workbook's date fixes before importing."
                )
            notes.append(f"row {r}: corrected date {expected} -> {corrected}")
            date = corrected
            applied_fixes += 1
        elif not corrections.covers(date):
            raise ValueError(
                f"Debug row {r}: date {date} is outside the "
                f"{corrections.tax_year}/{str(corrections.tax_year + 1)[-2:]} fiscal year. "
                "Add it to the workbook's date fixes after confirming the correct value."
            )

        account_from = _clean(ws.cell(r, 4).value)
        if r in corrections.accounts:
            expected, corrected = corrections.accounts[r]
            if str(account_from) != expected:
                raise ValueError(
                    f"Debug row {r}: expected account {expected!r}, found {account_from!r}. "
                    "The workbook has changed -- re-verify the workbook's account fixes."
                )
            notes.append(f"row {r}: account {expected!r} -> {corrected!r}")
            account_from = corrected
            applied_account_fixes += 1
        if account_from and str(account_from).strip().lower() == "cash":
            raise ValueError(
                f"Debug row {r}: cash transaction found. Cash is no longer tracked "
                "(DESIGN.md 2.1) and must not be imported."
            )

        amount = _dec(ws.cell(r, 10).value)
        if r in corrections.amounts:
            expected, corrected = corrections.amounts[r]
            if amount != expected:
                raise ValueError(
                    f"Debug row {r}: expected known-bad amount {expected}, found {amount}. "
                    "The workbook has changed -- re-verify the workbook's amount fixes before importing."
                )
            notes.append(
                f"row {r}: amount {expected} -> {corrected} (month tab is authoritative)"
            )
            amount = corrected
            applied_amount_fixes += 1

        category = _clean(ws.cell(r, 9).value)
        if r in corrections.categories:
            expected, corrected = corrections.categories[r]
            if category != expected:
                raise ValueError(
                    f"Debug row {r}: expected category {expected!r}, found {category!r}. "
                    "The workbook has changed -- re-verify the workbook's category fixes before importing."
                )
            notes.append(f"row {r}: category {expected!r} -> {corrected!r}")
            category = corrected
            applied_category_fixes += 1

        removed = bool(ws.cell(r, 15).value)
        removed_reason = None
        if r in corrections.phantoms:
            removed, removed_reason = True, corrections.phantoms[r]
            notes.append(f"row {r}: soft-deleted -- {removed_reason}")
            applied_phantoms += 1

        created = ws.cell(r, 14).value
        rows.append(
            LedgerRow(
                source_row=r,
                txn_date=date,
                identifier=str(ws.cell(r, 2).value),
                month=str(ws.cell(r, 3).value),
                account_from=str(account_from),
                account_to=_clean(ws.cell(r, 5).value),
                type=str(ws.cell(r, 8).value),
                category=category,
                amount=amount,
                category_comment=_clean(ws.cell(r, 11).value),
                comment=_clean(ws.cell(r, 12).value),
                classification=_clean(ws.cell(r, 13).value),
                created_at=created if isinstance(created, dt.datetime) else None,
                removed=removed,
                removed_reason=removed_reason,
            )
        )

    for label, applied, expected_total in (
        ("date", applied_fixes, len(corrections.dates)),
        ("amount", applied_amount_fixes, len(corrections.amounts)),
        ("category", applied_category_fixes, len(corrections.categories)),
        ("account", applied_account_fixes, len(corrections.accounts)),
        ("phantom", applied_phantoms, len(corrections.phantoms)),
    ):
        if applied != expected_total:
            raise ValueError(
                f"Expected {expected_total} {label} corrections, applied {applied}."
            )

    return rows, notes


# ---------------------------------------------------------------------- month tab shape


_SIGN_RE = re.compile(r"^=\$?[A-Z]{1,3}\$?\d+\s*([+-])")


def month_layout(values: Workbook, formulas: Workbook, month: str, period: str) -> MonthLayout:
    """Discover the account blocks and key rows of one month tab."""
    _, _, start_row, _, _ = resolve(values, f"xlStart{month}")
    _, _, end_row, _, _ = resolve(values, f"xlEnd{month}")

    vw, fw = values[month], formulas[month]
    layout = MonthLayout(name=month, period=period, start_row=start_row, end_row=end_row)

    col = FIRST_ACCOUNT_COL
    while True:
        name = _clean(vw.cell(2, col).value)
        if not name:
            break
        # Bank:        = <start> + <credit> - <debit>
        # Credit card: = <start> - <credit> + <debit>   (balance is positive debt)
        formula = fw.cell(4, col).value
        match = _SIGN_RE.match(str(formula or ""))
        if not match:
            raise ValueError(f"{month}: cannot determine type of account {name!r} from {formula!r}")
        acct_type = "bank" if match.group(1) == "+" else "credit_card"
        layout.blocks.append(AccountBlock(name=str(name), base_col=col, type=acct_type))
        col += COLS_PER_ACCOUNT

    return layout


def read_opening_balances(values: Workbook, layout: MonthLayout) -> dict[str, Decimal]:
    """Row 60 'Start' per account; the value sits two columns right of the label."""
    ws = values[layout.name]
    return {
        b.name: _dec(ws.cell(layout.start_row, b.base_col + 2).value) for b in layout.blocks
    }


def read_closing_balances(values: Workbook, layout: MonthLayout) -> dict[str, Decimal]:
    """Row 61 'End' per account -- the figure reconciliation must reproduce."""
    ws = values[layout.name]
    return {
        b.name: _dec(ws.cell(layout.end_row, b.base_col + 2).value) for b in layout.blocks
    }


def read_budgets(values: Workbook, month: str) -> dict[str, tuple[Decimal, Decimal]]:
    """Category -> (income, expected) from month-tab columns C and D."""
    sheet, col, min_row, _, max_row = resolve(values, f"xlCategories{month}")
    ws = values[sheet]
    out: dict[str, tuple[Decimal, Decimal]] = {}
    for r in range(min_row, max_row + 1):
        name = _clean(ws.cell(r, col).value)
        if not name:
            continue
        out[str(name)] = (_dec(ws.cell(r, col + 1).value), _dec(ws.cell(r, col + 2).value))
    return out


def read_projections(values: Workbook) -> list[tuple[dt.date, str, Decimal, str | None]]:
    """Projected Costs: (date, classification, amount, comment) per populated day.

    Column headers give the classifications, so a renamed or added one is picked up without
    touching this. The sheet holds a single month at a time -- whichever was last worked on.
    """
    ws = values["Projected Costs"]
    _, date_col, first_row, _, last_row = resolve(values, "xlProjDates")

    columns: dict[int, str] = {}
    for col in range(date_col + 3, date_col + 11):  # E..K, between Month and Total
        header = _clean(ws.cell(7, col).value)
        if header and str(header) != "Total":
            columns[col] = str(header)

    out = []
    for row in range(first_row, last_row + 1):
        raw = ws.cell(row, date_col).value
        if raw is None:
            continue
        day = raw.date() if isinstance(raw, dt.datetime) else raw
        comment = _clean(ws.cell(row, date_col + 11).value)  # M: Comments
        for col, name in columns.items():
            out.append((day, name, _dec(ws.cell(row, col).value), comment))
    return out


def read_daily_allowance(values: Workbook, month: str) -> Decimal:
    """'Spend per day' from the month tab's summary block.

    Found by its label rather than by cell reference: April, May and June carry extra
    inserted rows, so the value sits at C55 there and C54 from July onwards. Reading C54
    everywhere silently returns nothing for the first three months.
    """
    ws = values[month]
    for row in range(45, 62):
        label = _clean(ws.cell(row, 2).value)
        if label and str(label).strip().lower() == "spend per day":
            return _dec(ws.cell(row, 3).value)
    return Decimal("0")


_TRAILING_CONSTANT = re.compile(r"([+-]\s*\d+(?:\.\d+)?)\s*$")


def read_running_opening(formulas: Workbook, values: Workbook, month: str) -> dict[str, Decimal]:
    """Opening adjustments hardcoded onto a month's running-total formulas.

    April's Running Excess ends in a literal '-2632.45' -- the balance brought forward from
    the previous year's workbook, typed straight into the formula because there was nowhere
    else to put it. Recovered rather than lost.
    """
    sheet, min_col, _, max_col, _ = resolve(values, f"xlClass{month}")
    fws, vws = formulas[sheet], values[sheet]

    out: dict[str, Decimal] = {}
    for col in range(min_col, max_col + 1):
        header = vws.cell(1, col).value
        if not header or not str(header).startswith("Running "):
            continue
        match = _TRAILING_CONSTANT.search(str(fws.cell(4, col).value or ""))
        if match:
            out[str(header)[len("Running "):]] = Decimal(match.group(1).replace(" ", ""))
    return out


# Salary tracker bands -> assumption keys. Column B names the section, C the band, D holds
# the monthly figure.
#
# Keyed by (section, band) rather than by row, because the rows move: 25-26 models no
# additional rate at all, so its 'Adjusted bands' block sits one row higher than 26-27's
# and a row-based read takes the adjusted personal allowance of 475.83 for a 45% rate --
# storing it as 47,583%. The section is needed as well as the band because 'Personal
# allowance', 'Basic rate' and 'Higher rate' each appear twice, once raw and once adjusted.
SALARY_BANDS: dict[tuple[str, str], str] = {
    ("NI", "LEL"): "ni_lower_earnings_limit",
    ("NI", "UEL"): "ni_upper_earnings_limit",
    ("NI", "Lower rate"): "ni_lower_rate",
    ("NI", "Higher rate"): "ni_higher_rate",
    ("PAYE", "Personal allowance"): "personal_allowance",
    ("PAYE", "Basic rate threshold"): "basic_rate_threshold",
    ("PAYE", "Higher rate threshold"): "higher_threshold",
    ("PAYE", "Basic rate"): "basic_rate",
    ("PAYE", "Higher rate"): "higher_rate",
    ("PAYE", "Additional rate"): "additional_rate",
    ("Adjusted bands", "Basic rate"): "basic_band",  # D28 - D22
}

# The personal allowance is tapered in steps, each with its own start date in column F.
_PA_STEP = re.compile(r"^PA\s*-\s*\d+$")

# Held as percentages rather than fractions -- see models.SalaryAssumption. The workbook
# stores them as fractions, so they are scaled on the way in.
RATE_KEYS = frozenset(
    {"ni_lower_rate", "ni_higher_rate", "basic_rate", "higher_rate", "additional_rate"}
)


def read_salary_assumptions(
    values: Workbook, tax_year: int
) -> list[tuple[str, dt.date, Decimal]]:
    """(key, effective_from, monthly value) for the PAYE/NI bands.

    Everything except the personal-allowance steps applies from the start of the year; the
    steps carry their own start dates alongside them.

    A band the workbook does not model is simply absent from the result rather than read as
    zero -- 25-26 has no additional rate, and a 0% additional rate is a different claim from
    'not stated'.
    """
    ws = values["Salary tracker"]
    year_start = dt.date(tax_year, 4, 1)

    out = []
    section = None
    for row in range(17, ws.max_row + 1):
        heading = _clean(ws.cell(row, 2).value)  # B: NI | PAYE | Adjusted bands
        if heading:
            section = str(heading)
        band = _clean(ws.cell(row, 3).value)  # C
        if not band or section is None:
            continue
        band = str(band)

        if _PA_STEP.match(band):
            raw = ws.cell(row, 6).value  # F: effective from
            if raw is not None:
                start = raw.date() if isinstance(raw, dt.datetime) else raw
                out.append(
                    ("personal_allowance_adjustment", start, _dec(ws.cell(row, 4).value))
                )
            continue

        key = SALARY_BANDS.get((section, band))
        if key is None:
            continue
        value = _dec(ws.cell(row, 4).value)  # D: the monthly figure
        out.append((key, year_start, value * 100 if key in RATE_KEYS else value))

    return out


def read_payslips(values: Workbook, ref: RefData) -> list[dict]:
    """Salary tracker rows 4..15: actual payslip alongside the expected inputs."""
    ws = values["Salary tracker"]
    rows = []
    for row in range(4, 16):
        month = _clean(ws.cell(row, 2).value)
        if not month:
            continue
        rows.append(
            {
                "period": ref.period_for(str(month)),
                "payday": int(ws.cell(row, 3).value) if ws.cell(row, 3).value else None,
                "gross": _dec(ws.cell(row, 4).value) or None,
                "ni": _dec(ws.cell(row, 5).value) or None,
                "holiday_pay": _dec(ws.cell(row, 7).value) or None,
                "cycle_to_work": _dec(ws.cell(row, 9).value) or None,
                "paye": _dec(ws.cell(row, 11).value) or None,
                "net": _dec(ws.cell(row, 13).value) or None,
                "salary": _dec(ws.cell(row, 15).value) or None,      # O: annual salary
                "expected_gross": _dec(ws.cell(row, 16).value) or None,  # P
                "benefits": _dec(ws.cell(row, 17).value) or None,   # Q: benefits
                "additional": _dec(ws.cell(row, 18).value) or None,  # R: additional pay
            }
        )
    return rows


# A bonus month's annual salary written as its two parts: '=126022.4+7854'.
_SALARY_PLUS_BONUS = re.compile(r"^=\s*(-?\d+(?:\.\d+)?)\s*\+\s*(-?\d+(?:\.\d+)?)\s*$")


def read_salary_extras(
    formulas: Workbook, values: Workbook, ref: RefData
) -> dict[str, tuple[Decimal | None, Decimal]]:
    """period -> (annual salary, bonus), read from the formulas rather than the values.

    A bonus is never a field of its own in the workbook; it is welded into whichever cell
    was convenient, and the two workbooks chose differently:

        26-27 May   P5 = ROUND(O5/12,2)+29028.48        O holds the salary
        25-26 May   O5 = 126022.4+7854                  O holds salary *plus* bonus
                    P5 = ROUND((O5-7854)/12,2)+7854

    Column P's trailing constant is the bonus in both. What differs is column O: read as a
    value, 25-26's May says the annual salary was 133,876.40, which turns one month's bonus
    into a pay rise that reverses the month after -- and leaves a bonus of 7,199.50, a
    figure that appears nowhere and that nobody was ever paid.

    Deriving the bonus as expected_gross - salary/12 gets 26-27 right and 25-26 wrong, so
    both come from the formulas here.
    """
    fws, vws = formulas["Salary tracker"], values["Salary tracker"]

    out: dict[str, tuple[Decimal | None, Decimal]] = {}
    for row in range(4, 16):
        month = _clean(vws.cell(row, 2).value)
        if not month:
            continue
        match = _TRAILING_CONSTANT.search(str(fws.cell(row, 16).value or ""))  # P
        bonus = Decimal(match.group(1).replace(" ", "")) if match else Decimal("0")

        salary = _dec(vws.cell(row, 15).value) or None  # O
        split = _SALARY_PLUS_BONUS.match(str(fws.cell(row, 15).value or ""))
        if split and bonus and Decimal(split.group(2)) == bonus:
            salary = Decimal(split.group(1))

        out[ref.period_for(str(month))] = (salary, bonus)
    return out


def read_cards(values: Workbook) -> list[dict]:
    """Balance Transfer Cards: the parameter block, and each card's opening from the
    schedule beside it.

    Neither is at a fixed column. The schedule carries two columns per card, so the
    parameter block starts wherever the last card leaves off -- P3 in 26-27 with five
    cards, N3 in 25-26 with four. Reading 26-27's column against 25-26 lands on 'Term
    (months)' and names a card '10'.
    """
    ws = values["Balance Transfer Cards"]

    # Balance columns start at C and repeat every two columns (balance, payment). A card
    # taken out mid-year is zero until it starts, so the opening is the first non-zero row
    # rather than April's -- Halifax and MBNA 2 both begin in June.
    balances: dict[str, tuple[Decimal, dt.date]] = {}
    for col in range(3, ws.max_column + 1, 2):
        name = _clean(ws.cell(2, col).value)
        if not name or str(name) == "Total":
            continue
        for row in range(4, ws.max_row + 1):
            amount = _dec(ws.cell(row, col).value)
            raw = ws.cell(row, 2).value
            if amount and raw is not None:
                start = raw.date() if isinstance(raw, dt.datetime) else raw
                balances[str(name)] = (amount, start)
                break

    # 'Payment dates' heads the first parameter column; the card names are the column
    # before it, and term and minimum payment the two after.
    name_col = None
    for col in range(3, ws.max_column + 1):
        if str(ws.cell(2, col).value or "").strip() == "Payment dates":
            name_col = col - 1
            break
    if name_col is None:
        raise ValueError("Balance Transfer Cards: no 'Payment dates' heading in row 2")

    cards = []
    for order, row in enumerate(range(3, ws.max_row + 1)):
        name = _clean(ws.cell(row, name_col).value)
        if not name:
            break  # the block is contiguous; the first gap ends it
        # A card that never carries a balance still gets a sensible start date rather than
        # a sentinel: Tesco is zero throughout.
        first_row = ws.cell(4, 2).value
        default_start = (
            first_row.date() if isinstance(first_row, dt.datetime) else first_row
        ) or dt.date(1900, 1, 1)
        opening, start = balances.get(str(name), (Decimal("0"), default_start))
        cards.append(
            {
                "name": str(name),
                "opening_balance": opening,
                "opening_date": start,
                "payment_day": (
                    int(ws.cell(row, name_col + 1).value)
                    if ws.cell(row, name_col + 1).value
                    else None
                ),
                "term_months": int(_dec(ws.cell(row, name_col + 2).value)),
                "min_payment_pct": _dec(ws.cell(row, name_col + 3).value),
                "display_order": order,
            }
        )
    return cards


def read_cycling(values: Workbook) -> tuple[list[dict], list[dict], dict[str, Decimal]]:
    """Cycling: (outgoings, days ridden, saving rates).

    The rates are recovered from the L3 formula -- IF(commute, 10.5, IF(band, 8.9,
    IF(gym, 4.6, 0))) -- so they become data rather than staying inside a formula.
    """
    ws = values["Cycling"]

    outgoings = []
    for row in range(3, ws.max_row + 1):
        raw = ws.cell(row, 2).value
        if raw is None:
            continue
        day = raw.date() if isinstance(raw, dt.datetime) else raw
        outgoings.append(
            {
                "date": day,
                "item": _clean(ws.cell(row, 3).value),
                "amount": _dec(ws.cell(row, 4).value),
                "flag": _clean(ws.cell(row, 5).value),
            }
        )

    days = []
    rates = {"commute": Decimal("0"), "band": Decimal("0"), "gym": Decimal("0")}
    for row in range(3, ws.max_row + 1):
        raw = ws.cell(row, 8).value
        if raw is None:
            continue
        day = raw.date() if isinstance(raw, dt.datetime) else raw
        flags = {
            "commute": str(ws.cell(row, 9).value or "").upper() == "Y",
            "band": str(ws.cell(row, 10).value or "").upper() == "Y",
            "gym": str(ws.cell(row, 11).value or "").upper() == "Y",
        }
        days.append({"date": day, **flags})

        # L holds the saving the nested IF produced. Read the rate off the first day where
        # exactly one flag is set, so each rate comes from an unambiguous row.
        amount = _dec(ws.cell(row, 12).value)
        set_flags = [k for k, on in flags.items() if on]
        if amount and len(set_flags) == 1 and not rates[set_flags[0]]:
            rates[set_flags[0]] = amount

    return outgoings, days, rates


def summary_matrix(values: Workbook, ref: RefData) -> tuple[int, dict[int, str], range]:
    """Locate the Summary sheet's month-by-classification block.

    Returns (month column, {column: classification}, rows).

    Not a fixed reference, because the block floats. It sits to the right of the credit
    card table, and 25-26 has two cards fewer than 26-27 -- so the same block is Q19:Y31 in
    one workbook and O19:W31 in the other. Reading 26-27's coordinates against 25-26 does
    not fail cleanly either: it lands two columns left, in the card table, and asks
    period_for() about a month named '-4123.37'.

    Found by content instead. Row 19 holds two 'Month' headers, one per table; the one this
    wants is followed by classification names, which the other is not.
    """
    ws = values["Summary"]
    known = {c.name for c in ref.classifications}

    for col in range(1, 40):
        if str(ws.cell(19, col).value or "").strip() != "Month":
            continue
        headers = {}
        for right in range(col + 1, col + 10):
            label = _clean(ws.cell(19, right).value)
            if not label:
                break
            headers[right] = str(label)
        if len(known & set(headers.values())) < 2:
            continue  # the credit card table: Barclaycard, Halifax, Tesco, Payment
        rows = [r for r in range(20, 40) if _clean(ws.cell(r, col).value) in ref.months]
        return col, headers, range(rows[0], rows[-1] + 1)

    raise ValueError("Summary: no month-by-classification block found in row 19")


def read_daily_class_totals(
    values: Workbook, month: str, classification: str
) -> dict[dt.date, Decimal]:
    """Per-day total for one classification, from xlTotal<Class><Month>."""
    d_sheet, d_col, d_min, _, d_max = resolve(values, f"xlCatDates{month}")
    t_sheet, t_col, t_min, _, t_max = resolve(values, f"xlTotal{classification}{month}")
    dws, tws = values[d_sheet], values[t_sheet]

    out: dict[dt.date, Decimal] = {}
    for offset in range(min(d_max - d_min, t_max - t_min) + 1):
        raw = dws.cell(d_min + offset, d_col).value
        if raw is None:
            continue
        day = raw.date() if isinstance(raw, dt.datetime) else raw
        out[day] = _dec(tws.cell(t_min + offset, t_col).value)
    return out
